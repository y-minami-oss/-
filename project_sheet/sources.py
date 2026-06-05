"""データの取得元(データソース)。

2通りの読み取り方を用意する:

1. ライブ(gspread): 本番。Google スプレッドシートを直接読む。Mac で運用する想定。
2. オフライン(openpyxl): スプレッドシートを .xlsx でダウンロードしたファイルを読む。
   OAuth 設定なしで動作確認・テストができる。

どちらも「列名 -> 値」の辞書リストに整えて返すので、集計ロジック(aggregate.py)は
データソースを意識しない。
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from . import dc


# --------------------------------------------------------------------------
# オフライン: ダウンロード済み .xlsx を読む
# --------------------------------------------------------------------------
def read_dc_rows_from_xlsx(path: str | Path, sheet_name: str = dc.DEFAULT_SHEET_NAME) -> list[dict[str, Any]]:
    """.xlsx の「簡単DC」タブを読み、ヘッダー名をキーにした辞書リストで返す。"""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"タブ '{sheet_name}' が {path} に見つかりません。タブ一覧: {wb.sheetnames}")
    ws = wb[sheet_name]
    return _rows_from_grid([[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                            for r in range(1, ws.max_row + 1)])


# --------------------------------------------------------------------------
# ライブ: gspread で Google スプレッドシートを読む
# --------------------------------------------------------------------------
def read_dc_rows_from_gsheet(client: Any, spreadsheet_id: str,
                             sheet_name: str = dc.DEFAULT_SHEET_NAME) -> list[dict[str, Any]]:
    """gspread クライアントで効果測定シートの「簡単DC」タブを読む。"""
    sh = client.open_by_key(spreadsheet_id)
    ws = sh.worksheet(sheet_name)
    return _rows_from_grid(ws.get_all_values())


# --------------------------------------------------------------------------
# 共通: 2次元グリッド(1行目=ヘッダー) -> 辞書リスト
# --------------------------------------------------------------------------
def _rows_from_grid(grid: list[list[Any]]) -> list[dict[str, Any]]:
    if not grid:
        return []
    header = [str(h).strip() if h is not None else "" for h in grid[0]]
    rows: list[dict[str, Any]] = []
    for raw in grid[1:]:
        row = {header[i]: raw[i] for i in range(min(len(header), len(raw)))}
        rows.append(row)
    return rows
