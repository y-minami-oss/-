"""マネジメントシート(全体シート)から、担当者のアクティブ案件を抽出する。

全体シートのヘッダー行(「クライアント名」を含む行)を自動で見つけ、
- クライアント名 / No.
- 担当者名(CS/SSP/TSP/CC のそれぞれ)
- ステータス(アクティブ/終了済/停止中…)
- PJシート名
を読み取る。owner(既定:「三並」)がいずれかの担当に入っていて、かつ
ステータスが「アクティブ」の行だけを対象に返す。
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

ROLE_HEADERS = ("CS担当", "SSP担当", "TSP担当", "CC担当")


@dataclass
class ClientRecord:
    no: str
    client: str
    assignees: dict[str, str] = field(default_factory=dict)  # 役割 -> 担当者名
    status: str = ""
    pj_sheet: str = ""

    @property
    def is_active(self) -> bool:
        return "アクティブ" in (self.status or "")

    def roles_of(self, person: str) -> list[str]:
        return [role for role, name in self.assignees.items() if name == person]


def _find_header_row(grid: list[list[Any]]) -> int:
    for i, row in enumerate(grid):
        if any(str(c).strip() == "クライアント名" for c in row):
            return i
    raise ValueError("ヘッダー行(「クライアント名」を含む行)が見つかりません。")


def parse_master_grid(grid: list[list[Any]]) -> list[ClientRecord]:
    """全体シートの2次元グリッドから ClientRecord 一覧を作る。"""
    h = _find_header_row(grid)
    header = [str(c).strip() if c is not None else "" for c in grid[h]]

    def col(name: str) -> int | None:
        return header.index(name) if name in header else None

    c_client, c_no = col("クライアント名"), col("No.")
    c_status, c_pj = col("ステータス"), col("PJシート")
    role_cols = {r: col(r) for r in ROLE_HEADERS}

    records: list[ClientRecord] = []
    for raw in grid[h + 1:]:
        def get(idx: int | None) -> Any:
            return raw[idx] if (idx is not None and idx < len(raw)) else None

        client = get(c_client)
        if not client or not str(client).strip():
            continue
        assignees = {
            role: str(get(idx) or "").strip()
            for role, idx in role_cols.items()
            if str(get(idx) or "").strip()
        }
        records.append(
            ClientRecord(
                no=str(get(c_no) or "").strip(),
                client=str(client).strip(),
                assignees=assignees,
                status=str(get(c_status) or "").strip(),
                pj_sheet=str(get(c_pj) or "").strip(),
            )
        )
    return records


def active_clients_for(records: list[ClientRecord], owner: str) -> list[ClientRecord]:
    """owner(担当者名)がいずれかの担当に入っていて、アクティブな案件。"""
    return [r for r in records if r.is_active and owner in r.assignees.values()]


# --- 読み取りアダプタ -------------------------------------------------------
def read_master_grid_from_xlsx(path: str | Path, sheet_name: str = "全体シート") -> list[list[Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]


def read_master_grid_from_gsheet(client: Any, spreadsheet_id: str, sheet_name: str = "全体シート") -> list[list[Any]]:
    sh = client.open_by_key(spreadsheet_id)
    ws = sh.worksheet(sheet_name)
    return ws.get_all_values()
